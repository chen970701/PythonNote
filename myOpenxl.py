import openpyxl
import os
from openpyxl.styles import Font,PatternFill,Border,Alignment
from openpyxl.cell import cell
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import RGB
from openpyxl.chart.reference import Reference
from openpyxl.chart.bar_chart import BarChart
from openpyxl.chart.bar_chart import BarChart3D
import random
from openpyxl.chart.area_chart import AreaChart
#xls 快捷键ctrl + ; ctrl+shif+; 可以弹出时间
#-------------读xlsx----------------
#myworkbook=openpyxl.load_workbook(os.path.join(os.getcwd(),"GreenTestXL.xlsx"))
# print(myworkbook.sheetnames)
# mySheet1=myworkbook.worksheets[1] #sheet1
# #获取单元格值的方法
# #print(mySheet1["A1"].value)
# for v in range(1,11):
#     print(mySheet1.cell(v,1).value)
# myworkbook.close() 

#mySheet3=myworkbook.worksheets[2]
#print("该表格的为{}行{}列".format(mySheet3.max_row,mySheet3.max_column))
#遍列表格
# for r in range(mySheet3.max_row):
#     for c in range(mySheet3.max_column):
#         print(mySheet3.cell(r+1,c+1).value)
#mySheet3.title="奖牌榜"
#遍历第一列的所有值
#for i in range(len(list(mySheet3.columns)[0])):
#    print(list(mySheet3.columns)[0][i].value)
#myworkbook.save("GreenTestXL.xlsx")


#------------写Xlsx-------------
# myXSL=openpyxl.Workbook()
# sheetA=myXSL.create_sheet("sheetA",index=0)
# sheetC=myXSL.create_sheet("sheetC")
# sheetB=myXSL.create_sheet("sheetB",1)
# groups=sheetA["A1":"D5"]
# for lines in groups:
#     for cell in lines:
#         cell.value=cell.coordinate
# myXSL.save("save.xlsx")

# myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
# sheet1=myXSL.worksheets[0]

# sheet1.title="样式"
# print(dir(sheet1["B1"]))

# sheet1["B1"]="字体"
# sheet1["B1"].font=Font(name="华文琥珀",italic=True,color="FF0000")

# sheet1["B2"]="背景"
# sheet1["B2"].font=Font(name="楷体")
# sheet1["B2"].fill=PatternFill("solid",fgColor="00f000")
# # print(type(sheet1["B1"]))
# myXSL.save("GreenTestXL.xlsx")
# myXSL.close()



# values =Reference(sheet1,min_col=2,min_row=2,max_col=4,max_row=5)
# chart=BarChart()
# chart.add_data(values)
# chart.title="奖牌榜"
# sheet1.add_chart(chart,"E1")
# myXSL.save("GreenTestXL.xlsx")
# myXSL.close()


# myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
# # sheet1=myXSL.create_sheet("股票")
# # sheet1["A1"]="日期"
# # sheet1["B1"]="创业板"
# # sheet1["C1"]="大盘指数"
# # for t in range(1,31):
# #     sheet1.cell(t+1,1,str(t)+"日")
# #     dapan=round(random.random()*10,2)
# #     sheet1.cell(t+1,2,dapan+round(random.random()*5,2)-2)
# #     sheet1.cell(t+1,3,dapan)
# sheet1=myXSL.worksheets[3]
# chart=AreaChart()
# datas=Reference(sheet1,min_col=2,max_col=3,min_row=1,max_row=31)
# chart.add_data(datas,titles_from_data=True)#title从第一个数据生成
# sheet1.add_chart(chart,"E2")
# chart.x_axis.title="日期"
# chart.y_axis.title="股票价格"
# xdes=Reference(sheet1,min_col=1,min_row=2,max_row=31)
# chart.set_categories(xdes)
# myXSL.save("GreenTestXL.xlsx")
# myXSL.close()

#柱状图
# myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
# sheet1=myXSL.worksheets[2]
# chart=BarChart()
# chart.title="奖牌榜"
# chart.x_axis.title="国家"
# chart.y_axis.title="奖牌数"
# datas=Reference(sheet1,min_col=2,max_col=4,min_row=1,max_row=5)
# chart.add_data(datas,titles_from_data=True)#title从第一个数据生成
# xdes=Reference(sheet1,min_col=1,min_row=2,max_row=5)
# chart.set_categories(xdes)
# sheet1.add_chart(chart,"F2") 
# myXSL.save("GreenTestXL.xlsx")
# myXSL.close()

#3D柱状图
# myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
# sheet1=myXSL.worksheets[2]
# chart=BarChart3D()
# chart.title="奖牌榜3D"
# chart.x_axis.title="国家"
# chart.y_axis.title="奖牌数"
# datas=Reference(sheet1,min_col=2,max_col=4,min_row=1,max_row=5)
# chart.add_data(datas,titles_from_data=True)#title从第一个数据生成
# xdes=Reference(sheet1,min_col=1,min_row=2,max_row=5)
# chart.set_categories(xdes)
# sheet1.add_chart(chart,"F18") 
# myXSL.save("GreenTestXL.xlsx")
# myXSL.close()

#设置样式
# myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
# sheet1=myXSL.worksheets[2]
# sheet1["A1"]="统计"
# sheet1["A1"].font=Font(name="华文琥珀",italic=True,color="0000FF")
# sheet1["A1"].fill=PatternFill("solid",fgColor="aaa89d")
# for lines in sheet1["B1:D1"]:
#     for onecell in lines:
#         onecell.fill=PatternFill("solid",fgColor="fef2b8")
# for lines in sheet1["A1:A5"]:
#     for onecell in lines:
#         onecell.font=Font(name="黑体")
#         onecell.fill=PatternFill("solid",fgColor="f8a5e2")
# sheet1.row_dimensions[1].height=20 #第一行
# sheet1.column_dimensions["A"].width=10 #第A列
# sheet1.freeze_panes="B2"

myXSL=openpyxl.load_workbook("GreenTestXL.xlsx")
sheet1=myXSL.worksheets[2]
sheet1["A1"]="统计"
sheet1["A1"].font=Font(name="华文琥珀",italic=True,color="0000FF")
sheet1["A1"].fill=PatternFill("solid",fgColor="aaa89d")
for lines in sheet1["B1:D1"]:
    for onecell in lines:
        onecell.fill=PatternFill("solid",fgColor="fef2b8")
for lines in sheet1["A1:A5"]:
    for onecell in lines:
        onecell.font=Font(name="黑体")
        onecell.fill=PatternFill("solid",fgColor="f8a5e2")
sheet1.row_dimensions[1].height=20 #第一行
sheet1.column_dimensions["A"].width=10 #第A列
sheet1.freeze_panes="B2"
